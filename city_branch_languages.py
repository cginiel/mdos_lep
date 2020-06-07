from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import json
import secrets # file that contains your API key

###################### global vars ######################
MAPQUEST_KEY = secrets.CONSUMER_KEY
RESOURCE_URL = "http://www.mapquestapi.com/search/v2/radius?"

CACHE_FILENAME = "michigan_LEP_cache.json"
CACHE_DICT = {}

mdos_addresses_xlsx = "mdos-building-addresses.xlsx"
lep_by_county_xlsx = "lep-by-county-michigan.xlsx"

county_list = []
county_primary_foreign_lang_dict = {}
county_secondary_foreign_lang_dict = {}
#########################################################


###################### caching ##########################
def open_cache():
    ''' Opens the cache file if it exists and loads the JSON into
    the CACHE_DICT dictionary.
    if the cache file doesn't exist, creates a new cache dictionary
    
    Parameters
    ----------
    None
    
    Returns
    -------
    dict
        the opened cache
    '''
    try:
        cache_file = open(CACHE_FILENAME, 'r')
        cache_contents = cache_file.read()
        cache_dict = json.loads(cache_contents)
        cache_file.close()
    except:
        cache_dict = {}
    return cache_dict


def save_cache(cache_dict):
    ''' Saves the current state of the cache to disk
    
    Parameters
    ----------
    cache_dict: dict
        The dictionary to save
    
    Returns
    -------
    None
    '''
    dumped_json_cache = json.dumps(cache_dict)
    fw = open(CACHE_FILENAME,"w")
    fw.write(dumped_json_cache)
    fw.close() 


def make_request_with_cache(key, value):
    '''Issues a request to the cache saved to the device.

    If the item exists in the cache, the program will pull from that data.
    If the item does not exist in the cache, the program will create a key/value
    pair and save it to the cache dictionary

    Parameters
    ----------
    key
        list
    value
        list

    '''
    if key in CACHE_DICT.keys():
        print("Using cache")
        return CACHE_DICT[key]

    else:
        print("Fetching")
        CACHE_DICT[key] = value
        save_cache(CACHE_DICT)
        return CACHE_DICT[key]

################################################

def import_workbook(filename):
    '''imports an excel workbook as a file python can work with.

    params
    ------
    filename : str
        string that points to .xlsx file

    returns
    -------
    wb : object
        excel workbook object
    '''
    wb = load_workbook(filename)

    return wb


def get_mdos_building_zipcodes():
    '''reads the mdos-building-addresses.xlsx file and extracts zipcodes that can
    eventually be sent through the mapquest API in order to find out which county
    each zipcode is in.

    params
    ------
    none

    returns
    -------
    zipcode_list : list
        list of zipcodes as they correspond to the row on the excel sheet. 
        e.g., C2 = 49221; C145 = 48202
    '''
    zipcode_list = []
    ## importing our workbook
    branches = import_workbook(mdos_addresses_xlsx)

    ## accessing a specific worksheet
    address_sheet = branches['Address']

    ## access zipcode from each address cell
    for address in address_sheet['C2':'C145']:
        full_address = address[0].value

        ## some zipcodes have the extension so I cleaned that off
        if "-" in full_address[-10:]:
            zipcode = full_address[-10:-5]
            zipcode_list.append(zipcode)
        else:
            zipcode = full_address[-5:]
            zipcode_list.append(zipcode)

    return zipcode_list
    

def search_for_county_with_zipcode(zipcode):
    '''searches a zipcode on the mapquest API and returns a corresponding county.

    params
    ------
    zipcode : int

    returns
    -------
    county : str
        the county associated with the zipcode provided
    '''
    params = {
    "key" : MAPQUEST_KEY,
    "origin" : zipcode,
    "radius" : 10,
    "maxMatches" : 10,
    "ambiguities" : "ignore",
    "outFormat" : "json"
    }

    param_strings = []
    connector = '&'
    for k in params.keys():
        param_strings.append(f'{k}={params[k]}')
    param_strings.sort()
    unique_key = RESOURCE_URL + connector.join(param_strings)
    response = requests.get(unique_key).json()
    county = response['origin']['adminArea4']
    make_request_with_cache(zipcode, county)

    return county


def make_county_list_from_zipcode():
    '''Uses the county/zipcode match from search_for_county_with_zipcode() to create a county list
    that matches each MDOS building from the mdos-building-addresses.xlsx file.

    params
    ------
    none

    returns
    -------
    county_list : list
        list of counties as they correspond to the mdos-building-addresses.xlsx rows.
    '''
    for zipcode in get_mdos_building_zipcodes():
        county = search_for_county_with_zipcode(zipcode)
        if county == "":
            county = "Saginaw County"
            county_list.append(county)
        else:
            county_list.append(county)

    return county_list


def apply_column_header_styles(cell):
    '''Applies a blue fill to a column header's cell

    params
    ------
    cell : object
        an xlsx cell to update with the blue fill

    returns
    ------
    none
    '''
    ## add header styles
    blueFill = PatternFill(
        start_color = 'FFB4C6E7',
        end_color = 'FFB4C6E7',
        fill_type = 'solid')

    cell.fill = blueFill


def add_counties_to_mdos_branches():
    '''takes the county list and adds a new column of counties to be matched with the corresponding MDOS branch.
    Saves an excel file with the updated column of counties.

    params
    ------
    none

    returns
    -------
    none
        mdos-building-addresses.xlsx : saves an excel file with counties that match MDOS locations
    '''

    ## import our workbook
    branches = import_workbook(mdos_addresses_xlsx)
    address_sheet = branches['Address']

    county_column = address_sheet['D2':'D145']

    ## this is where we update the spreadsheet based on county 
    for i in (range(len(county_column))):
        county_column[i][0].value = county_list[i]

    ## add blue fill
    apply_column_header_styles(address_sheet['D1'])

    ## add header text
    address_sheet['D1'].value = 'County'

    ## save our changes to the excel sheet
    branches.save("mdos-building-addresses.xlsx")


def make_primary_foreign_lang_by_county_dict():
    '''Creates a dictionary of LEP info based on county.
    We can use this dictionary to add information to the MDOS branches excel sheet.

    params
    ------
    none

    returns
    -------
    lep_by_county_dict : dict
        a dictionary where counties are the key and corresponding LEP info are the values.
    '''
    lep_county_list = []
    lep_primary_foreign_language_list = []

    ## importing excel workbook
    lep_info = import_workbook(lep_by_county_xlsx)

    ## open sheet from workbook
    lep_by_county_sheet = lep_info['County']

    ## make a county list (these will be keys of our dict)
    lep_county_column = lep_by_county_sheet['A6':'A88']

    for cell in lep_county_column:
        lep_county_list.append(cell[0].value)

    ## make a language list (these will be the counties' values)
    lep_language_column = lep_by_county_sheet['D6':'D88']

    for cell in lep_language_column:
        if cell[0].value == " ":
            cell[0].value = "No language reported"
            lep_primary_foreign_language_list.append(cell[0].value)
        else:
            lep_primary_foreign_language_list.append(cell[0].value)

    ## make the dict
    for i in range(len(lep_county_list)):
        county_primary_foreign_lang_dict[lep_county_list[i]] = lep_primary_foreign_language_list[i]

    return county_primary_foreign_lang_dict


def make_secondary_foreign_lang_by_county_dict():
    '''Creates a dictionary of LEP info based on county.
    We can use this dictionary to add information to the MDOS branches excel sheet.

    params
    ------
    none

    returns
    -------
    lep_by_county_dict : dict
        a dictionary where counties are the key and corresponding LEP info are the values.
    '''
    lep_county_list = []
    lep_secondary_foreign_language_list = []

    ## importing excel workbook
    lep_info = import_workbook(lep_by_county_xlsx)

    ## open sheet from workbook
    lep_by_county_sheet = lep_info['County']

    ## make a county list (these will be keys of our dict)
    lep_county_column = lep_by_county_sheet['A6':'A88']

    for cell in lep_county_column:
        lep_county_list.append(cell[0].value)

    ## make a language list (these will be the counties' values)
    lep_language_column = lep_by_county_sheet['G6':'G88']

    for cell in lep_language_column:
        if cell[0].value == " ":
            cell[0].value = "No language reported"
            lep_secondary_foreign_language_list.append(cell[0].value)
        else:
            lep_secondary_foreign_language_list.append(cell[0].value)

    ## make the dict
    for i in range(len(lep_county_list)):
        county_secondary_foreign_lang_dict[lep_county_list[i]] = lep_secondary_foreign_language_list[i]

    return county_secondary_foreign_lang_dict


def add_foreign_lang_info_to_mdos_branches():
    '''adds LEP primary and secondary foreign languages from the lep-by-county excel file to the 
    mdos-building-addresses excel file.

    params
    ------
    none

    returns
    -------
    none
        mdos-building-addresses-with-county-and-primary-foreign-lang.xlsx : saves an excel file with counties that match MDOS locations
    '''
    ## importing our excel sheet
    branches = import_workbook(mdos_addresses_xlsx)
    address_sheet = branches['Address']

    ## assign variable to the column to which we wish to apply our languages
    ## primary foreign lang
    primary_foreign_language_column = address_sheet['E2':'E145']

    ## secondary foreign lang
    secondary_foreign_language_column = address_sheet['F2':'F145']

    ## match languages to their respective county
    for i in range(len(county_list)):

        ## add primary foreign lang
        if county_list[i] in county_primary_foreign_lang_dict:
            primary_foreign_language_column[i][0].value = county_primary_foreign_lang_dict[county_list[i]]

        ## add secondary foreign lang
        if county_list[i] in county_secondary_foreign_lang_dict:
            secondary_foreign_language_column[i][0].value = county_secondary_foreign_lang_dict[county_list[i]]

    ## add primary foreign lang header styles and text
    apply_column_header_styles(address_sheet['E1'])
    ## update title text
    address_sheet['E1'].value = "Primary Foreign Language"

    ## add secondary foreign lang header styles and text
    apply_column_header_styles(address_sheet['F1'])
    ## update title text
    address_sheet['F1'].value = "Secondary Foreign Language"

    branches.save("mdos-building-addresses-with-county-and-foreign-languages.xlsx")


if __name__ == "__main__":
    CACHE_DICT = open_cache()
    make_county_list_from_zipcode()
    add_counties_to_mdos_branches()
    make_primary_foreign_lang_by_county_dict()
    make_secondary_foreign_lang_by_county_dict()
    add_foreign_lang_info_to_mdos_branches()


